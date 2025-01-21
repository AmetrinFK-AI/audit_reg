import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font

# Заголовок застосунку
st.title("Аудит регіону")

# Інструкція для користувача
st.write("""
Цей застосунок дозволяє:
- Завантажити Excel-файл з даними про лоти (очікуються листи "Sheet1" та "Sheet2").
- Відфільтрувати дані за вказаними компаніями.
- Показати **всі області** за різними показниками.
- Завантажити результат у форматі XLSX з автопідбором ширини стовпців.
- **Додатково**:
  - Створити окремий лист для кожної області з топ-20 компаніями за сумою виграних тендерів, 
    де три вказані компанії об'єднані в одну "АМЕТРІН ФК".
  - Додати колонку "2023" на основі даних з "Sheet2" з колонки "Поточна сума договорів лота".
  - Перейменувати колонку "Сума виграних тендерів" на "2024" в листах областей.
  - Додати колонку "динаміка" (%) за формулою ((2024-2023)/2023)*100 з округленням до цілих та символом "%".
  - У першому рядку регіональних листів показати загальні суми за 2024 (з Sheet1) та 2023 (з Sheet2) по області.
  - Додати колонки "Доля 2024" та "Доля 2023", які розраховуються відносно значень у першому рядку.
  - **Змінено**: Колонка "Прирост долі" рахується як ((Доля 2024 / Доля 2023) - 1)*100%, якщо Доля 2023 = 0, то приріст = 0.
  - **Нове**: Створити окремі листи "Доля Клиента" для кожної області.
    - Виділити жирним текстом назви організаторів.
    - Об’єднати три вказані компанії в одну "АМЕТРІН ФК".
""")

# Завантаження файлу
uploaded_file = st.file_uploader("Завантажте Excel-файл з даними", type=["xlsx"])

# Оригінальні компанії для групування
original_target_companies = [
    'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "ДІЯ ФАРМ"',
    'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "АМЕТРІН ФК"',
    'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "МОДЕРН-ФАРМ"'
]

grouped_company_name = "АМЕТРІН ФК"
companies_to_group = original_target_companies

if uploaded_file:
    try:
        # Перевірка та зчитування листів
        data_sheet1 = pd.read_excel(
            uploaded_file,
            sheet_name='Sheet1',
            usecols=['Поточна сума договорів лота', 'Переможець', 'Регіон організатора', 'Організатор']
        )
        data_sheet2 = pd.read_excel(
            uploaded_file,
            sheet_name='Sheet2',
            usecols=['Поточна сума договорів лота', 'Переможець', 'Регіон організатора']
        )

        # Перевірка обов’язкових стовпців у Sheet1
        required_columns_sheet1 = [
            'Поточна сума договорів лота',
            'Переможець',
            'Регіон організатора',
            'Організатор'
        ]
        missing_cols_sheet1 = [col for col in required_columns_sheet1 if col not in data_sheet1.columns]
        if missing_cols_sheet1:
            st.error(f"Відсутні необхідні стовпці у Sheet1: {', '.join(missing_cols_sheet1)}")
            st.stop()

        # Перевірка обов’язкових стовпців у Sheet2
        required_columns_sheet2 = [
            'Поточна сума договорів лота',
            'Переможець',
            'Регіон організатора'
        ]
        missing_cols_sheet2 = [col for col in required_columns_sheet2 if col not in data_sheet2.columns]
        if missing_cols_sheet2:
            st.error(f"Відсутні необхідні стовпці у Sheet2: {', '.join(missing_cols_sheet2)}")
            st.stop()

        # Обробка Sheet1
        data_sheet1['Поточна сума договорів лота'] = (
            data_sheet1['Поточна сума договорів лота']
            .astype(str)
            .str.replace('\u00a0', '', regex=True)   # Видаляємо нерозривні пробіли
            .str.replace(',', '.', regex=True)       # Заміна коми на крапку (за потреби)
            .replace('-', pd.NA)                     # Заміна '-' на NaN
        )
        data_sheet1['Поточна сума договорів лота'] = pd.to_numeric(
            data_sheet1['Поточна сума договорів лота'], errors='coerce'
        )

        # Обробка Sheet2
        data_sheet2['Поточна сума договорів лота'] = (
            data_sheet2['Поточна сума договорів лота']
            .astype(str)
            .str.replace('\u00a0', '', regex=True)
            .str.replace(',', '.', regex=True)
            .replace('-', pd.NA)
        )
        data_sheet2['Поточна сума договорів лота'] = pd.to_numeric(
            data_sheet2['Поточна сума договорів лота'], errors='coerce'
        )

        # Приводимо назву "Переможець" до базового вигляду (зрізаємо все після "|")
        data_sheet1['Переможець'] = (
            data_sheet1['Переможець']
            .astype(str)
            .str.split('|').str[0]
            .str.strip()
        )
        # Об'єднуємо 3 компанії в "АМЕТРІН ФК"
        data_sheet1['Переможець'] = data_sheet1['Переможець'].apply(
            lambda x: grouped_company_name if x in companies_to_group else x
        )
        # Видаляємо рядки, де Переможець = "-"
        data_sheet1 = data_sheet1[data_sheet1['Переможець'] != '-']

        # Те саме для Sheet2
        data_sheet2['Переможець'] = (
            data_sheet2['Переможець']
            .astype(str)
            .str.split('|').str[0]
            .str.strip()
        )
        data_sheet2['Переможець'] = data_sheet2['Переможець'].apply(
            lambda x: grouped_company_name if x in companies_to_group else x
        )
        data_sheet2 = data_sheet2[data_sheet2['Переможець'] != '-']

        # -------------------------------
        # Логіка формування підсумкових таблиць
        # -------------------------------

        # Загальні суми по регіонах за 2024 (Sheet1)
        total_region_summary = (
            data_sheet1
            .dropna(subset=['Регіон організатора'])
            [data_sheet1['Регіон організатора'] != '-']
            .groupby('Регіон організатора', as_index=False)
            .agg(total_sum=('Поточна сума договорів лота', 'sum'))
        )

        # Загальні суми по регіонах за 2023 (Sheet2)
        total_sum_dict_2023 = (
            data_sheet2
            .groupby('Регіон організатора')['Поточна сума договорів лота']
            .sum()
            .to_dict()
        )

        # Фільтрація для згрупованої компанії (2024)
        filtered_data = data_sheet1[
            (data_sheet1['Переможець'] == grouped_company_name) &
            data_sheet1['Регіон організатора'].notna() &
            (data_sheet1['Регіон організатора'] != '-')
        ]

        if filtered_data.empty:
            st.warning("За вказаними компаніями не знайдено жодного лота.")
            st.stop()

        # Підрахунок суми та кількості тендерів для компанії (2024)
        companies_region_summary = (
            filtered_data
            .groupby('Регіон організатора', as_index=False)
            .agg(
                sum_companies=('Поточна сума договорів лота', 'sum'),
                count_companies=('Поточна сума договорів лота', 'count')
            )
        )

        # Мерджимо з загальними сумами по регіонах
        merged_summary = pd.merge(
            total_region_summary,
            companies_region_summary,
            on='Регіон організатора',
            how='inner'
        )

        # Перейменування стовпців
        merged_summary.rename(columns={
            'Регіон організатора': 'Область',
            'total_sum': '2024',
            'sum_companies': 'Сума виграних тендерів компаній',
            'count_companies': 'Кількість виграних тендерів'
        }, inplace=True)

        # Додаємо 2023
        merged_summary['2023'] = merged_summary['Область'].apply(
            lambda x: total_sum_dict_2023.get(x, 0)
        )

        # Додаємо колонку "доля"
        merged_summary['доля'] = (
            merged_summary['Сума виграних тендерів компаній'] / merged_summary['2024'] * 100
        ).round(2).astype(str) + '%'

        # Раніше тут було обмеження top-20, тепер виводимо всі області
        merged_summary = merged_summary.sort_values(by='Сума виграних тендерів компаній', ascending=False)

        st.subheader("Усі області")
        st.dataframe(merged_summary, use_container_width=True)

        # Підготовка даних для детальних листів
        sum_2023_grouped = data_sheet2.groupby(
            ['Регіон організатора', 'Переможець']
        )['Поточна сума договорів лота'].sum().reset_index()

        sum_2023_dict = {}
        for _, row in sum_2023_grouped.iterrows():
            region = row['Регіон організатора']
            company = row['Переможець']
            sum_2023_dict.setdefault(region, {})[company] = row['Поточна сума договорів лота']

        total_sum_dict_2024 = merged_summary.set_index('Область')['2024'].to_dict()

        # Створюємо Excel у пам’яті
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Записуємо основну зведену таблицю
            merged_summary.to_excel(writer, sheet_name='Data', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Автопідбір ширини стовпців
            for i, col in enumerate(merged_summary.columns, start=1):
                max_length = max(
                    merged_summary[col].astype(str).map(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=i).column_letter
                ].width = max_length + 2

            # Функції для розрахунків
            def calc_dynamic(row):
                if row['2023'] == 0:
                    return '0%'
                return f"{round(((row['2024'] - row['2023']) / row['2023']) * 100)}%"

            def calc_share(val, total):
                if total == 0 or pd.isna(val):
                    return 0
                return (val / total) * 100

            def calc_growth(share_2024, share_2023):
                if share_2023 == 0:
                    return 0
                return (share_2024 / share_2023 - 1) * 100

            # Створюємо листи по кожному регіону з топ-20 компаніями
            for region in merged_summary['Область']:
                region_data = data_sheet1[data_sheet1['Регіон організатора'] == region]

                top_companies = (
                    region_data
                    .groupby('Переможець', as_index=False)
                    .agg(total_sum=('Поточна сума договорів лота', 'sum'))
                    .sort_values(by='total_sum', ascending=False)
                    .head(20)  # Тут залишаємо обмеження топ-20 компаній
                )

                top_companies.rename(columns={
                    'Переможець': 'Назва компанії',
                    'total_sum': '2024'
                }, inplace=True)

                # Додаємо 2023
                top_companies['2023'] = top_companies.apply(
                    lambda row: sum_2023_dict.get(region, {}).get(row['Назва компанії'], 0),
                    axis=1
                )

                # Динаміка
                top_companies['динаміка'] = top_companies.apply(calc_dynamic, axis=1)

                total_sum_region_2024 = total_sum_dict_2024.get(region, 0)
                total_sum_region_2023 = total_sum_dict_2023.get(region, 0)

                # Рядок "ВСЬОГО"
                summary_row = pd.DataFrame([{
                    'Назва компанії': 'ВСЬОГО',
                    '2024': total_sum_region_2024,
                    '2023': total_sum_region_2023,
                    'динаміка': ''
                }])

                top_companies = pd.concat([summary_row, top_companies], ignore_index=True)

                # Розрахунок долі
                top_companies['share_2024_num'] = top_companies['2024'].apply(
                    lambda x: calc_share(x, total_sum_region_2024)
                )
                top_companies['share_2023_num'] = top_companies['2023'].apply(
                    lambda x: calc_share(x, total_sum_region_2023)
                )

                # Прирост долі
                top_companies['Прирост долі'] = top_companies.apply(
                    lambda row: calc_growth(row['share_2024_num'], row['share_2023_num']),
                    axis=1
                )

                # Форматуємо у відсотки
                top_companies['Доля 2024'] = top_companies['share_2024_num'].round(0).astype(int).astype(str) + '%'
                top_companies['Доля 2023'] = top_companies['share_2023_num'].round(0).astype(int).astype(str) + '%'
                top_companies['Прирост долі'] = top_companies['Прирост долі'].round(0).astype(int).astype(str) + '%'

                # Фінальна структура
                top_companies = top_companies[[
                    'Назва компанії', '2024', '2023', 'динаміка',
                    'Доля 2024', 'Доля 2023', 'Прирост долі'
                ]]

                # Назва листа для регіону
                base_sheet_name = f"2024_{region}"
                sheet_name = base_sheet_name[:31]
                original_sheet_name = sheet_name
                counter = 1
                while sheet_name in writer.book.sheetnames:
                    suffix = f"_{counter}"
                    sheet_name = f"{original_sheet_name[:31 - len(suffix)]}{suffix}"
                    counter += 1
                    if counter > 100:
                        st.error(f"Неможливо створити унікальне ім'я для листа області: {region}")
                        break

                top_companies.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # Автопідбір ширини стовпців
                for i, col in enumerate(top_companies.columns, start=1):
                    max_length = max(
                        top_companies[col].astype(str).map(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[
                        worksheet.cell(row=1, column=i).column_letter
                    ].width = max_length + 2

            # Створюємо листи "Доля Клиента" для кожного регіону
            for region in merged_summary['Область']:
                region_data = data_sheet1[data_sheet1['Регіон організатора'] == region]

                top_20_organizers_region = (
                    region_data
                    .groupby('Організатор', as_index=False)
                    .agg(total_lot=('Поточна сума договорів лота', 'sum'))
                    .sort_values(by='total_lot', ascending=False)
                    .head(20)
                )

                доля_кліента_rows = []
                for _, organizer_row in top_20_organizers_region.iterrows():
                    organizer = organizer_row['Організатор']
                    organizer_sum = organizer_row['total_lot']

                    # Строка з організатором (жирний шрифт)
                    доля_кліента_rows.append({
                        'Організатор/Переможець': organizer,
                        'Сума лота': organizer_sum,
                        'Доля (%)': ''
                    })

                    # Знаходимо переможців для цього організатора
                    winners = (
                        region_data[region_data['Організатор'] == organizer]
                        .groupby('Переможець', as_index=False)
                        .agg(total_lot=('Поточна сума договорів лота', 'sum'))
                        .sort_values(by='total_lot', ascending=False)
                    )

                    # Якщо немає "АМЕТРІН ФК", додаємо з 0
                    if not any(winners['Переможець'] == grouped_company_name):
                        winners = pd.concat([
                            winners,
                            pd.DataFrame({
                                'Переможець': [grouped_company_name],
                                'total_lot': [0]
                            })
                        ], ignore_index=True)

                    for _, winner_row in winners.iterrows():
                        winner = winner_row['Переможець']
                        winner_sum = winner_row['total_lot']
                        share = (winner_sum / organizer_sum * 100) if organizer_sum != 0 else 0

                        доля_кліента_rows.append({
                            'Організатор/Переможець': f"    {winner}",
                            'Сума лота': winner_sum,
                            'Доля (%)': f"{share:.2f}%"
                        })

                доля_кліента_df = pd.DataFrame(доля_кліента_rows)

                base_sheet_name = f"Доля Клиента {region}"
                sheet_name = base_sheet_name[:31]
                original_sheet_name = sheet_name
                counter = 1
                while sheet_name in writer.book.sheetnames:
                    suffix = f"_{counter}"
                    sheet_name = f"{original_sheet_name[:31 - len(suffix)]}{suffix}"
                    counter += 1
                    if counter > 100:
                        st.error(f"Неможливо створити унікальне ім'я для листа 'Доля Клиента' регіону: {region}")
                        break

                доля_кліента_df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # Автопідбір ширини стовпців
                for i, col in enumerate(доля_кліента_df.columns, start=1):
                    max_length = max(
                        доля_кліента_df[col].astype(str).map(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[
                        worksheet.cell(row=1, column=i).column_letter
                    ].width = max_length + 2

                # Жирний шрифт для "Організаторів" (рядки без відступу)
                bold_font = Font(bold=True)
                for row in worksheet.iter_rows(
                        min_row=2,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=1
                ):
                    for cell in row:
                        if not str(cell.value).startswith('    '):
                            cell.font = bold_font

                # Заголовки теж робимо жирними
                for cell in worksheet["1:1"]:
                    cell.font = bold_font

            # Закриваємо writer
            writer.close()

            # Отримуємо байти з пам'яті
            processed_data = output.getvalue()

            if len(processed_data) == 0:
                st.error("Помилка: Оброблений файл порожній.")
            else:
                st.success("Обробку даних завершено успішно!")
                st.download_button(
                    label="Завантажити оброблений Excel-файл",
                    data=processed_data,
                    file_name="Аудит_регіону.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Помилка обробки даних: {e}")
else:
    st.info("Будь ласка, завантажте Excel-файл для початку аналізу.")
